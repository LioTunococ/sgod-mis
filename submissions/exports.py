from __future__ import annotations

import csv
import io
from dataclasses import dataclass
import json
from typing import Any, Iterable, List

from django.utils.text import slugify

from . import constants as smea_constants


@dataclass(slots=True)
class ExportTable:
    """Container describing a logical table in an export."""

    title: str
    headers: List[str]
    rows: List[List[Any]]

    def sheet_name(self) -> str:
        """Return a workbook-safe sheet name (<= 31 chars)."""
        base = slugify(self.title) or "sheet"
        return base[:31]


@dataclass(slots=True)
class SubmissionExport:
    filename_prefix: str
    tables: List[ExportTable]

    def iter_tables(self) -> Iterable[ExportTable]:
        return tuple(self.tables)


def _build_filename_prefix(submission) -> str:
    school_code = slugify(submission.school.code) if submission.school else "submission"
    period_label = slugify(submission.period.label) if submission.period else "period"
    return f"{school_code}-{submission.id}-{period_label}"[:64]




def _grade_numbers_for_submission(submission):
    school = getattr(submission, "school", None)
    if not school:
        return sorted(smea_constants.GRADE_NUMBER_TO_LABEL.keys())
    profile = getattr(school, "profile", None)
    start = getattr(profile, "grade_span_start", None)
    end = getattr(profile, "grade_span_end", None)
    if start is None or end is None:
        start = getattr(school, "min_grade", None)
        end = getattr(school, "max_grade", None)
    if start is None or end is None:
        return sorted(smea_constants.GRADE_NUMBER_TO_LABEL.keys())
    start = max(0, start)
    end = max(start, end)
    return [n for n in range(start, end + 1) if n in smea_constants.GRADE_NUMBER_TO_LABEL]


def _slp_pair_index(submission):
    pairs = []
    for grade_number in _grade_numbers_for_submission(submission):
        grade_label = smea_constants.GRADE_NUMBER_TO_LABEL[grade_number]
        subjects = smea_constants.SLP_SUBJECTS_BY_GRADE.get(grade_number, [smea_constants.SLP_DEFAULT_SUBJECT])
        for subject_code, _ in subjects:
            pairs.append((grade_label, subject_code))
    if not pairs:
        pairs.append((smea_constants.GRADE_NUMBER_TO_LABEL.get(0, "Grade 1"), smea_constants.SLP_DEFAULT_SUBJECT[0]))
    return {pair: index for index, pair in enumerate(pairs)}

def _build_school_profile_table(submission) -> ExportTable:
    school = getattr(submission, "school", None)
    profile = getattr(school, "profile", None) if school else None
    strands = ", ".join(profile.strands) if getattr(profile, "strands", None) else ""
    grade_span = school.grade_span_label if school else ""
    district_name = getattr(getattr(school, "district", None), "name", "") if school else ""
    rows = [
        ["School", getattr(school, "name", "")],
        ["District", district_name],
        ["School Head", getattr(profile, "head_name", "") if profile else ""],
        ["Contact", getattr(profile, "head_contact", "") if profile else ""],
        ["Grade Span", grade_span],
        ["Strands / Programs", strands],
    ]
    return ExportTable(
        title="School Profile",
        headers=["Field", "Value"],
        rows=rows,
    )


def build_slp_export(submission) -> SubmissionExport:
    pair_index = _slp_pair_index(submission)
    slp_rows_sorted = sorted(
        submission.form1_slp_rows.all(),
        key=lambda row: pair_index.get((row.grade_label, row.subject), len(pair_index)),
    )
    def _summarize_interventions(value: Any) -> str:
        text = value or ""
        if not isinstance(text, str):
            return str(text)
        t = text.strip()
        if not t:
            return ""
        # If stored as JSON array of {code, reason, intervention}
        try:
            data = json.loads(t)
            if isinstance(data, list):
                parts = []
                for i, item in enumerate(data, start=1):
                    if not isinstance(item, dict):
                        continue
                    reason = (item.get("reason") or item.get("code") or "").strip()
                    interv = (item.get("intervention") or "").strip()
                    if reason or interv:
                        if interv:
                            parts.append(f"{i}. {reason}: {interv}")
                        else:
                            parts.append(f"{i}. {reason}")
                return "; ".join(parts) if parts else ""
        except Exception:
            pass
        # Fallback: original free text
        return t

    rows = [
        [
            row.grade_label,
            row.enrolment if row.is_offered else "",
            row.get_subject_display(),
            "Yes" if row.is_offered else "Not offered",
            row.dnme if row.is_offered else "",
            row.fs if row.is_offered else "",
            row.s if row.is_offered else "",
            row.vs if row.is_offered else "",
            row.o if row.is_offered else "",
            row.top_three_llc,
            getattr(row, 'non_mastery_reasons', ''),
            getattr(row, 'non_mastery_other', ''),
            _summarize_interventions(row.intervention_plan),
        ]
        for row in slp_rows_sorted
    ]

    analysis = getattr(submission, "form1_slp_analysis", None)
    analysis_row = [
        getattr(analysis, "q1a_summary_text", ""),
        getattr(analysis, "root_causes", ""),
        getattr(analysis, "best_practices", ""),
    ]

    dnme_rows = [
        [entry.position, entry.grade_label, entry.count]
        for entry in submission.form1_slp_top_dnme.all().order_by("position")
    ]
    outstanding_rows = [
        [entry.position, entry.grade_label, entry.count]
        for entry in submission.form1_slp_top_outstanding.all().order_by("position")
    ]

    tables = [
        ExportTable(
            title="SLP Learner Progress",
            headers=[
                "Grade",
                "Enrolment",
                "Subject",
                "Offered",
                "DNME",
                "FS",
                "S",
                "VS",
                "O",
                "Top 3 LLC",
                "Reasons (Codes)",
                "Other Reasons",
                "Intervention Plan",
            ],
            rows=rows,
        ),
        ExportTable(
            title="SLP Analysis",
            headers=["Q1A Summary", "Root Causes", "Best Practices"],
            rows=[analysis_row],
        ),
        ExportTable(
            title="SLP Top 5 DNME",
            headers=["Rank", "Grade", "Count"],
            rows=dnme_rows,
        ),
        ExportTable(
            title="SLP Top 5 Outstanding",
            headers=["Rank", "Grade", "Count"],
            rows=outstanding_rows,
        ),
    ]
    tables.insert(0, _build_school_profile_table(submission))
    return SubmissionExport(filename_prefix=_build_filename_prefix(submission), tables=tables)


def build_reading_export(submission) -> SubmissionExport:
    # Derive timing phrase from quarter
    q = getattr(getattr(submission, 'period', None), 'quarter_tag', None)
    def _timing_from_quarter(qtag: str | None) -> str:
        if qtag == 'Q1':
            return 'eosy'
        if qtag in ('Q2', 'Q3'):
            return 'bosy'
        if qtag == 'Q4':
            return 'mosy'
        return 'eosy'
    def _reading_phrase_for_timing(t: str) -> str:
        t = (t or '').lower()
        if t == 'bosy':
            return 'BOSY Results - Quarter 2 or 3 Depends on what quarter was selected'
        if t == 'mosy':
            return 'MOSY Results - Quarter 4'
        return 'EOSY Results - Quarter 1'
    derived_timing = _timing_from_quarter(q)
    timing_phrase = _reading_phrase_for_timing(derived_timing)

    crla_rows = [
        [
            entry.get_level_display(),
            entry.get_timing_display(),
            entry.get_subject_display(),
            entry.get_band_display(),
            entry.count,
        ]
        for entry in submission.form1_crla.all().order_by("level", "timing", "subject", "band")
    ]

    philiri_rows = [
        [
            entry.get_level_display(),
            entry.get_timing_display(),
            entry.get_language_display(),
            entry.band_4_7,
            entry.band_5_8,
            entry.band_6_9,
            entry.band_10,
        ]
        for entry in submission.form1_philiri.all().order_by("level", "timing", "language")
    ]

    interventions = [
        [entry.order, entry.description]
        for entry in submission.form1_reading_interventions.all().order_by("order")
    ]

    # Reading Difficulties & Interventions (structured)
    difficulties_rows: list[list[str]] = []
    # Prefer model rows if populated; fall back to JSON stored in submission.data
    try:
        rdp_qs = submission.reading_difficulty_plans.all().order_by('period', 'grade_label')
        for plan in rdp_qs:
            period = (plan.period or '').upper()
            grade = plan.get_grade_label_display() if hasattr(plan, 'get_grade_label_display') else plan.grade_label
            pairs = plan.data if isinstance(plan.data, list) else []
            for idx, pair in enumerate(pairs, start=1):
                diff = (pair.get('difficulty') or '').strip()
                interv = (pair.get('intervention') or '').strip()
                if diff or interv:
                    difficulties_rows.append([period, grade, idx, diff, interv])
    except Exception:
        difficulties_rows = []
    if not difficulties_rows:
        # Fallback: parse raw JSON structure [{grade, pairs:[{difficulty, intervention}]}]
        raw = (submission.data or {}).get('reading_difficulties_json', [])
        if isinstance(raw, list):
            for entry in raw:
                grade = str(entry.get('grade', '')).strip()
                pairs = entry.get('pairs') or []
                for idx, pair in enumerate(pairs, start=1):
                    diff = (pair.get('difficulty') or '').strip()
                    interv = (pair.get('intervention') or '').strip()
                    if diff or interv:
                        difficulties_rows.append(['', grade, idx, diff, interv])

    tables = [
        ExportTable(
            title="Reading Context",
            headers=["Note"],
            rows=[[timing_phrase]],
        ),
        ExportTable(
            title="Reading CRLA",
            headers=["Grade", "Timing", "Subject", "Band", "Learner Count"],
            rows=crla_rows,
        ),
        ExportTable(
            title="Reading PHILIRI",
            headers=["Grade", "Timing", "Language", "Band 4-7", "Band 5-8", "Band 6-9", "Band 10"],
            rows=philiri_rows,
        ),
        ExportTable(
            title="Reading Interventions",
            headers=["Order", "Description"],
            rows=interventions,
        ),
        ExportTable(
            title="Reading Difficulties & Targeted Interventions",
            headers=["Period", "Grade", "Pair #", "Difficulty", "Intervention"],
            rows=difficulties_rows,
        ),
    ]
    tables.insert(0, _build_school_profile_table(submission))
    return SubmissionExport(filename_prefix=_build_filename_prefix(submission), tables=tables)


def build_rma_export(submission) -> SubmissionExport:
    rma_rows = [
        [
            row.get_grade_label_display() if hasattr(row, "get_grade_label_display") else row.grade_label,
            row.enrolment,
            row.emerging_not_proficient,
            row.emerging_low_proficient,
            row.developing_nearly_proficient,
            row.transitioning_proficient,
            row.at_grade_level,
        ]
        for row in submission.form1_rma_rows.all().order_by("grade_label")
    ]

    rma_interventions = [
        [entry.order, entry.description]
        for entry in submission.form1_rma_interventions.all().order_by("order")
    ]

    tables = [
        ExportTable(
            title="RMA Results",
            headers=[
                "Grade",
                "Enrolment",
                "Not Proficient (<25%)",
                "Low (25-49%)",
                "Nearly Proficient (50-74%)",
                "Proficient (75-84%)",
                "At Grade Level (85%+)",
            ],
            rows=rma_rows,
        ),
        ExportTable(
            title="RMA Interventions",
            headers=["Order", "Description"],
            rows=rma_interventions,
        ),
    ]
    tables.insert(0, _build_school_profile_table(submission))
    return SubmissionExport(filename_prefix=_build_filename_prefix(submission), tables=tables)


def build_adm_export(submission) -> SubmissionExport:
    adm_rows = [
        [
            row.ppas_conducted,
            row.ppas_physical_target,
            row.ppas_physical_actual,
            row.ppas_physical_percent,
            row.funds_downloaded,
            row.funds_obligated,
            row.funds_unobligated,
            row.funds_percent_obligated,
            row.funds_percent_burn_rate,
            row.q1_response,
            row.q2_response,
            row.q3_response,
            row.q4_response,
            row.q5_response,
        ]
        for row in submission.form1_adm_rows.all().order_by("id")
    ]

    tables = [
        ExportTable(
            title="ADM Records",
            headers=[
                "PPAS Conducted",
                "Physical Target",
                "Physical Actual",
                "Physical %",
                "Funds Downloaded",
                "Funds Obligated",
                "Funds Unobligated",
                "% Obligated",
                "Burn Rate %",
                "Q1",
                "Q2",
                "Q3",
                "Q4",
                "Q5",
            ],
            rows=adm_rows,
        )
    ]
    tables.insert(0, _build_school_profile_table(submission))
    return SubmissionExport(filename_prefix=_build_filename_prefix(submission), tables=tables)


_TAB_BUILDERS = {
    "slp": build_slp_export,
    "reading": build_reading_export,
    "rma": build_rma_export,
    "adm": build_adm_export,
}


def build_export_for_tab(submission, tab: str) -> SubmissionExport:
    try:
        builder = _TAB_BUILDERS[tab]
    except KeyError as exc:
        raise ValueError(f"Unsupported export tab: {tab}") from exc
    return builder(submission)


def render_export_to_csv(export: SubmissionExport) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for idx, table in enumerate(export.iter_tables()):
        if idx:
            writer.writerow([])
        writer.writerow([f"# {table.title}"])
        writer.writerow(table.headers)
        for row in table.rows:
            writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def render_export_to_xlsx(export: SubmissionExport) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to render XLSX exports") from exc

    wb = Workbook()
    tables = list(export.iter_tables())
    if not tables:
        wb.create_sheet(title="Data")
    for index, table in enumerate(tables):
        if index == 0:
            ws = wb.active
            ws.title = table.sheet_name()
        else:
            ws = wb.create_sheet(title=table.sheet_name())
        ws.append(table.headers)
        for row in table.rows:
            ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



