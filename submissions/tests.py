from __future__ import annotations

import io
import unittest

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

OPENPYXL_AVAILABLE = load_workbook is not None

from accounts.models import UserProfile
from organizations.models import District, School, SchoolProfile, Section
from submissions import constants as smea_constants
from submissions.models import (
    Form1ADMRow,
    Form1PctHeader,
    Form1RMARow,
    Form1RMAIntervention,
    Form1ReadingCRLA,
    Form1ReadingIntervention,
    Form1ReadingPHILIRI,
    Form1SLPAnalysis,
    Form1SLPRow,
    Form1SLPTopDNME,
    Form1SLPTopOutstanding,
    FormTemplate,
    Period,
    Submission,
    SubmissionAttachment,
    SubmissionTimeline,
    SMEAActivityRow,
    SMEAProject,
)
from submissions.views import slp_grade_labels_for_school
from submissions.forms import Form1SLPRowForm
from submissions.exports import build_slp_export


class SubmissionWorkflowTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.section = Section.objects.create(code="smme", name="School Management")
        self.district = District.objects.create(code="north", name="North District")
        self.school = School.objects.create(code="test-school", name="Test School", district=self.district)
        self.period = Period.objects.create(
            label="Q1",
            school_year_start=2025,
            quarter_tag='Q1',
            display_order=1,
            is_active=True,
        )
        self.form = FormTemplate.objects.create(
            section=self.section,
            code="smea-form-1",
            title="SMEA Form 1",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=timezone.now().date(),
            close_at=timezone.now().date(),
        )
        self.school_head = User.objects.create_user(
            username="schoolhead", email="head@example.com", password="password"
        )
        profile = UserProfile.objects.get(user=self.school_head)
        profile.school = self.school
        profile.save(update_fields=["school", "updated_at"])

        self.section_admin = User.objects.create_user(
            username="sectionadmin", email="admin@example.com", password="password"
        )
        section_profile = UserProfile.objects.get(user=self.section_admin)
        section_profile.section_admin_codes = [self.section.code]
        section_profile.save(update_fields=["section_admin_codes", "updated_at"])

        self.psds = User.objects.create_user(
            username="psds", email="psds@example.com", password="password"
        )
        psds_profile = UserProfile.objects.get(user=self.psds)
        psds_profile.districts.add(self.district)

        self.school.implements_adm = True
        self.school.save(update_fields=["implements_adm"])

        self.submission = Submission.objects.create(
            school=self.school,
            form_template=self.form,
            period=self.period,
        )
        self.project = SMEAProject.objects.create(
            submission=self.submission,
            project_title="Project Alpha",
            area_of_concern="Reading",
        )
        SMEAActivityRow.objects.create(
            project=self.project,
            activity="Conduct baseline study",
        )


        Form1SLPRow.objects.create(
            submission=self.submission,
            grade_label="Grade 3",
            enrolment=30,
            dnme=5,
            fs=5,
            s=10,
            vs=5,
            o=5,
            top_three_llc="Reading",
            intervention_plan="Conduct remedial classes",
        )
        Form1SLPAnalysis.objects.create(
            submission=self.submission,
            q1a_summary_text="Summary",
            root_causes="Root Causes",
            best_practices="Practices",
        )
        Form1SLPTopDNME.objects.create(
            submission=self.submission, position=1, grade_label="Grade 3", count=5
        )
        Form1SLPTopOutstanding.objects.create(
            submission=self.submission, position=1, grade_label="Grade 6", count=3
        )

        Form1ReadingCRLA.objects.create(
            submission=self.submission,
            level=smea_constants.CRLALevel.GRADE3,
            timing=smea_constants.CRLATiming.BOY,
            subject=smea_constants.CRLASubject.ENGLISH,
            band=smea_constants.CRLABand.INSTRUCTIONAL,
            count=12,
        )
        Form1ReadingPHILIRI.objects.create(
            submission=self.submission,
            level=smea_constants.CRLALevel.GRADE3,
            timing=smea_constants.AssessmentTiming.BOY,
            language=smea_constants.PHILIRILanguage.ENGLISH,
            band_4_7=1,
            band_5_8=1,
            band_6_9=1,
            band_10=1,
        )
        Form1ReadingIntervention.objects.create(
            submission=self.submission, order=1, description="Reading club"
        )

        Form1RMARow.objects.create(
            submission=self.submission,
            grade_label=smea_constants.RMAGradeLabel.GRADE3,
            enrolment=40,
            band_below_75=5,
            band_75_79=5,
            band_80_84=10,
            band_85_89=10,
            band_90_100=10,
        )
        Form1RMAIntervention.objects.create(
            submission=self.submission, order=1, description="Focused review"
        )

        Form1ADMRow.objects.create(
            submission=self.submission,
            ppas_conducted="Orientation",
            ppas_physical_target=10,
            ppas_physical_actual=8,
            ppas_physical_percent=80,
            funds_downloaded=1000,
            funds_obligated=800,
            funds_unobligated=200,
            funds_percent_obligated=80,
            funds_percent_burn_rate=60,
            q1_response="Q1",
            q2_response="Q2",
            q3_response="Q3",
            q4_response="Q4",
            q5_response="Q5",
        )

    def test_submission_transitions(self):
        self.assertTrue(self.submission.can_submit())
        self.submission.mark_submitted(self.school_head)
        self.assertEqual(self.submission.status, Submission.Status.SUBMITTED)
        self.assertIsNotNone(self.submission.submitted_at)

        self.submission.mark_returned(self.section_admin, "Please update the data.")
        self.assertEqual(self.submission.status, Submission.Status.RETURNED)
        self.assertIn("update", self.submission.returned_remarks)

        self.submission.mark_submitted(self.school_head)
        self.assertEqual(self.submission.status, Submission.Status.SUBMITTED)

        self.submission.mark_noted(self.section_admin, "Looks good.")
        self.assertEqual(self.submission.status, Submission.Status.NOTED)
        self.assertIn("Looks", self.submission.noted_remarks)

    def test_reading_timing_enforced_by_quarter(self):
        # Login as school head to access edit view
        self.client.login(username="schoolhead", password="password")
        # Attempt to override timing via GET param should be ignored
        url = reverse("edit_submission", args=[self.submission.id]) + "?tab=reading&reading_period=mosy"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        # For Q1, enforced timing should be BOSY
        ctx = resp.context
        self.assertIsNotNone(ctx)
        # Policy: Q1 -> EOSY
        self.assertEqual(ctx.get("selected_reading_period"), "eosy")

    def test_timeline_records_transitions(self):
        timeline = list(SubmissionTimeline.objects.filter(submission=self.submission).order_by("created_at"))
        self.assertEqual(len(timeline), 1)
        self.assertEqual(timeline[0].to_status, Submission.Status.DRAFT)
        self.assertEqual(timeline[0].remarks, "Submission created")
        self.assertIsNone(timeline[0].actor)

        self.submission.mark_submitted(self.school_head)
        self.submission.mark_returned(self.section_admin, "Please update the data.")
        self.submission.mark_submitted(self.school_head)
        self.submission.mark_noted(self.section_admin, "Looks good.")

        timeline = list(SubmissionTimeline.objects.filter(submission=self.submission).order_by("created_at"))
        transitions = [(entry.from_status or "", entry.to_status) for entry in timeline]
        self.assertEqual(
            transitions,
            [
                ("", Submission.Status.DRAFT),
                (Submission.Status.DRAFT, Submission.Status.SUBMITTED),
                (Submission.Status.SUBMITTED, Submission.Status.RETURNED),
                (Submission.Status.RETURNED, Submission.Status.SUBMITTED),
                (Submission.Status.SUBMITTED, Submission.Status.NOTED),
            ],
        )
        self.assertEqual(timeline[1].actor, self.school_head)
        self.assertEqual(timeline[2].actor, self.section_admin)
        self.assertEqual(timeline[-1].remarks, "Looks good.")

    def test_slp_reasons_other_validation(self):
        row = Form1SLPRow.objects.filter(submission=self.submission).first()
        # Build form data for this row using the form prefix used in views (simulate partial save)
        prefix = 'slp_rows-0'
        data = {
            f'{prefix}-grade_label': row.grade_label,
            f'{prefix}-subject': row.subject,
            f'{prefix}-enrolment': row.enrolment,
            f'{prefix}-dnme': row.dnme,
            f'{prefix}-fs': row.fs,
            f'{prefix}-s': row.s,
            f'{prefix}-vs': row.vs,
            f'{prefix}-o': row.o,
            f'{prefix}-is_offered': 'on',
            # reasons selects 'f' but doesn't provide other -> invalid
            f'{prefix}-reasons': ['f'],
            f'{prefix}-reason_other': '',
        }
        form = Form1SLPRowForm(data=data, instance=row, prefix=prefix)
        self.assertFalse(form.is_valid())
        self.assertIn('reason_other', form.errors)

        # Provide sufficient sentences (>=2) -> valid
        data[f'{prefix}-reason_other'] = 'There was a change in teacher assignment. Materials arrived late.'
        form = Form1SLPRowForm(data=data, instance=row, prefix=prefix)
        self.assertTrue(form.is_valid(), msg=f"Unexpected errors: {form.errors}")
        saved = form.save()
        self.assertEqual(saved.non_mastery_reasons, 'f')
        self.assertIn('Materials', saved.non_mastery_other)

    def test_slp_export_includes_reasons_columns(self):
        # Add a row with reasons to check export
        row = Form1SLPRow.objects.filter(submission=self.submission).first()
        row.non_mastery_reasons = 'a,c'
        row.non_mastery_other = 'Contextual factors explained.'
        row.save(update_fields=['non_mastery_reasons', 'non_mastery_other'])
        export = build_slp_export(self.submission)
        # Find the learner progress table
        table = next(t for t in export.iter_tables() if t.title == 'SLP Learner Progress')
        self.assertIn('Reasons (Codes)', table.headers)
        self.assertIn('Other Reasons', table.headers)
        # Ensure at least one row has our codes
        found = any('a,c' in (r[table.headers.index('Reasons (Codes)')] or '') for r in table.rows)
        self.assertTrue(found)

    def test_slp_export_summarizes_interventions_from_json(self):
        row = Form1SLPRow.objects.filter(submission=self.submission).first()
        # Simulate new UI storing JSON
        row.intervention_plan = (
            '[{"code":"a","reason":"Pre-requisite skills were not mastered","intervention":"Conduct LAC on prerequisites"},'
            ' {"code":"b","reason":"The identified LLC are difficult to teach","intervention":"Team-teaching and demo"}]'
        )
        row.save(update_fields=['intervention_plan'])
        export = build_slp_export(self.submission)
        table = next(t for t in export.iter_tables() if t.title == 'SLP Learner Progress')
        idx = table.headers.index('Intervention Plan')
        # At least one row should contain a numbered, human-readable summary
        has_summary = any(
            ('1.' in (r[idx] or '')) and ('2.' in (r[idx] or '')) and ('Conduct LAC' in (r[idx] or ''))
            for r in table.rows
        )
        self.assertTrue(has_summary)

    def test_attachment_validation_blocks_invalid_extension(self):
        invalid = SimpleUploadedFile("notes.txt", b"dummy", content_type="text/plain")
        attachment = SubmissionAttachment(
            submission=self.submission,
            file=invalid,
            original_name="notes.txt",
        )
        with self.assertRaises(ValidationError):
            attachment.full_clean()

    def test_school_head_can_start_submission(self):
        self.client.force_login(self.school_head)
        url = reverse("start_submission", args=[self.form.code, self.period.id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Submission.objects.filter(
                school=self.school,
                form_template=self.form,
                period=self.period,
            ).exists()
        )

    def test_open_forms_list_shows_forms(self):
        self.client.force_login(self.school_head)
        url = reverse("open_forms_list", args=[self.section.code])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.form.title)
        self.assertContains(response, self.form.code)

    def test_submit_requires_project_and_activity(self):
        other_form = FormTemplate.objects.create(
            section=self.section,
            code="smea-form-2",
            title="Secondary Project",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=self.form.open_at,
            close_at=self.form.close_at,
        )
        new_submission = Submission.objects.create(
            school=self.school,
            form_template=other_form,
            period=self.period,
        )
        with self.assertRaises(ValidationError):
            new_submission.mark_submitted(self.school_head)
        project = SMEAProject.objects.create(submission=new_submission, project_title="Project Beta")
        with self.assertRaises(ValidationError):
            new_submission.mark_submitted(self.school_head)
        SMEAActivityRow.objects.create(project=project, activity="Initial activity")
        new_submission.mark_submitted(self.school_head)
        self.assertEqual(new_submission.status, Submission.Status.SUBMITTED)

    def test_section_admin_queue_requires_role(self):
        self.client.force_login(self.school_head)
        response = self.client.get(reverse("review_queue", args=[self.section.code]))
        self.assertEqual(response.status_code, 403)

        self.client.force_login(self.section_admin)
        response = self.client.get(reverse("review_queue", args=[self.section.code]))
        self.assertEqual(response.status_code, 200)

    def test_psds_queue_is_limited_to_assigned_districts(self):
        self.client.force_login(self.psds)

        self.submission.status = Submission.Status.SUBMITTED
        self.submission.submitted_at = timezone.now()
        self.submission.save(update_fields=["status", "submitted_at", "updated_at"])
        in_scope_submission = self.submission

        other_district = District.objects.create(code="south", name="South District")
        other_school = School.objects.create(
            code="south-es",
            name="South ES",
            district=other_district,
        )
        out_of_scope_submission = Submission.objects.create(
            school=other_school,
            form_template=self.form,
            period=self.period,
            status=Submission.Status.SUBMITTED,
            submitted_at=timezone.now(),
        )

        response = self.client.get(reverse("review_queue", args=[self.section.code]))
        self.assertEqual(response.status_code, 200)

        submissions = list(response.context["submissions"])
        self.assertIn(in_scope_submission, submissions)
        self.assertNotIn(out_of_scope_submission, submissions)

    def test_review_detail_includes_timeline(self):
        self.submission.mark_submitted(self.school_head)
        self.client.force_login(self.section_admin)
        response = self.client.get(reverse("review_detail", args=[self.submission.id]))
        self.assertEqual(response.status_code, 200)
        timeline = list(response.context["timeline_entries"])
        self.assertGreaterEqual(len(timeline), 2)
        self.assertEqual(timeline[0].to_status, Submission.Status.SUBMITTED)

    def test_edit_submission_shows_timeline(self):
        self.submission.mark_submitted(self.school_head)
        self.submission.mark_returned(self.section_admin, "Needs updates")
        self.client.force_login(self.school_head)
        response = self.client.get(reverse("edit_submission", args=[self.submission.id]))
        self.assertEqual(response.status_code, 200)
        timeline_entries = response.context["timeline_entries"]
        self.assertGreaterEqual(len(timeline_entries), 2)
        statuses = [entry.to_status for entry in timeline_entries]
        self.assertIn(Submission.Status.RETURNED, statuses)

    def test_section_admin_can_open_readonly_tabs(self):
        self.client.force_login(self.section_admin)
        response = self.client.get(reverse("review_submission_tabs", args=[self.submission.id]))
        self.assertEqual(response.status_code, 200)
        self.assertIn("tabs", response.context)
        self.assertIn("timeline_entries", response.context)
        self.assertEqual(response.context["current_tab"], "projects")

    def test_psds_can_open_readonly_tabs(self):
        self.client.force_login(self.psds)
        response = self.client.get(reverse("review_submission_tabs", args=[self.submission.id]))
        self.assertEqual(response.status_code, 200)

    def test_readonly_tabs_rejects_school_head(self):
        self.client.force_login(self.school_head)
        response = self.client.get(reverse("review_submission_tabs", args=[self.submission.id]))
        self.assertEqual(response.status_code, 403)

    def test_export_csv_contains_slp_headers(self):
        SchoolProfile.objects.update_or_create(
            school=self.school,
            defaults={"head_name": "Jamie Santos", "grade_span_start": 1, "grade_span_end": 6, "strands": ["STEM"]},
        )
        self.client.force_login(self.section_admin)
        url = reverse("review_submission_export", args=[self.submission.id, "csv"])
        response = self.client.get(f"{url}?tab=slp")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        decoded = response.content.decode()
        self.assertIn("# School Profile", decoded)
        self.assertIn("Jamie Santos", decoded)
        lines = [line for line in decoded.splitlines() if line and not line.startswith("#")]
        self.assertIn("Grade,Enrolment,DNME,FS,S,VS,O,Top 3 LLC,Intervention Plan", lines)

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is required for XLSX export tests")
    def test_export_xlsx_includes_reading_sheet(self):
        self.client.force_login(self.section_admin)
        url = reverse("review_submission_export", args=[self.submission.id, "xlsx"])
        response = self.client.get(f"{url}?tab=reading")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("reading-crla", [name.lower() for name in workbook.sheetnames])
        sheet = workbook["reading-crla"]
        header = [cell.value for cell in sheet[1]]
        self.assertEqual(header, ["Grade", "Timing", "Subject", "Band", "Learner Count"])

    def test_export_requires_reviewer_role(self):
        self.client.force_login(self.school_head)
        url = reverse("review_submission_export", args=[self.submission.id, "csv"])
        response = self.client.get(f"{url}?tab=slp")
        self.assertEqual(response.status_code, 403)

    def _build_pct_payload(self):
        header = Form1PctHeader.objects.get(submission=self.submission)
        rows = list(header.rows.order_by('area'))
        data = {
            'tab': 'pct',
            'pct-TOTAL_FORMS': len(rows),
            'pct-INITIAL_FORMS': len(rows),
            'pct-MIN_NUM_FORMS': 0,
            'pct-MAX_NUM_FORMS': len(rows),
            'action': 'save',
        }
        for index, row in enumerate(rows):
            data[f'pct-{index}-id'] = row.id
            data[f'pct-{index}-area'] = row.area
            data[f'pct-{index}-percent'] = row.percent
            data[f'pct-{index}-action_points'] = row.action_points
        return data

    def test_review_queue_exposes_quick_stats(self):
        self.client.force_login(self.section_admin)
        pending_form = FormTemplate.objects.create(
            section=self.section,
            code="smea-form-pending",
            title="SMEA Form Pending",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=timezone.now().date(),
            close_at=timezone.now().date(),
        )
        returned_form = FormTemplate.objects.create(
            section=self.section,
            code="smea-form-returned",
            title="SMEA Form Returned",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=timezone.now().date(),
            close_at=timezone.now().date(),
        )
        noted_form = FormTemplate.objects.create(
            section=self.section,
            code="smea-form-noted",
            title="SMEA Form Noted",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=timezone.now().date(),
            close_at=timezone.now().date(),
        )

        pending_submission = Submission.objects.create(
            school=self.school,
            form_template=pending_form,
            period=self.period,
            status=Submission.Status.SUBMITTED,
            submitted_at=timezone.now(),
        )
        returned_submission = Submission.objects.create(
            school=self.school,
            form_template=returned_form,
            period=self.period,
            status=Submission.Status.RETURNED,
            returned_at=timezone.now(),
        )
        noted_submission = Submission.objects.create(
            school=self.school,
            form_template=noted_form,
            period=self.period,
            status=Submission.Status.NOTED,
            noted_at=timezone.now(),
        )

        url = reverse('review_queue', args=[self.section.code])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        stats = response.context['quick_stats']
        self.assertEqual(stats['pending_total'], 1)
        self.assertEqual(stats['submitted_today'], 1)
        self.assertEqual(stats['returned_total'], 1)
        self.assertEqual(stats['noted_this_week'], 1)
        self.assertContains(response, 'Awaiting review')

    def test_locked_submission_blocks_updates(self):
        url = reverse('edit_submission', args=[self.submission.id])
        self.client.force_login(self.school_head)
        self.client.get(f"{url}?tab=pct")
        self.submission.status = Submission.Status.SUBMITTED
        self.submission.save(update_fields=['status'])
        payload = self._build_pct_payload()
        response = self.client.post(url, payload)
        self.assertEqual(response.status_code, 403)

    def test_grade_rows_follow_school_span(self):
        self.school.min_grade = 3
        self.school.max_grade = 4
        self.school.save(update_fields=['min_grade', 'max_grade'])
        url = reverse('edit_submission', args=[self.submission.id])
        self.client.force_login(self.school_head)
        self.client.get(f"{url}?tab=slp")
        labels = list(Form1SLPRow.objects.filter(submission=self.submission).values_list('grade_label', flat=True))
        self.assertEqual(labels, ['Grade 3', 'Grade 4'])
        self.client.get(f"{url}?tab=rma")
        rma_labels = list(Form1RMARow.objects.filter(submission=self.submission).values_list('grade_label', flat=True))
        self.assertEqual(rma_labels, ['g3', 'g4'])

    def test_school_profile_overrides_grade_span_helpers(self):
        SchoolProfile.objects.create(
            school=self.school,
            grade_span_start=2,
            grade_span_end=5,
        )
        labels = slp_grade_labels_for_school(self.school)
        self.assertEqual(labels, ["Grade 2", "Grade 3", "Grade 4", "Grade 5"])

    # Removed bulk strand toggle test: School Profile governs SHS offerings; UI and handler deleted.
class FormTemplateManagementTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.section = Section.objects.create(code="smme", name="SMME")
        self.other_section = Section.objects.create(code="hrd", name="HRD")
        today = timezone.now().date()
        self.form = FormTemplate.objects.create(
            section=self.section,
            code="smme-q1",
            title="SMME Q1",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=today,
            close_at=today,
            is_active=True,
        )
        self.section_admin = User.objects.create_user(username="sectionadmin", password="pass")
        profile = UserProfile.objects.get(user=self.section_admin)
        profile.section_admin_codes = [self.section.code]
        profile.save(update_fields=["section_admin_codes", "updated_at"])

    def test_manage_forms_page_renders(self):
        self.client.force_login(self.section_admin)
        response = self.client.get(reverse("manage_section_forms"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Section Form Management")
        self.assertContains(response, self.form.title)

    def test_section_admin_can_create_form(self):
        self.client.force_login(self.section_admin)
        today = timezone.now().date()
        payload = {
            "action": "create",
            "section": self.section.id,
            "code": "smme-q2",
            "title": "SMME Q2",
            "period_type": FormTemplate.PeriodType.QUARTER,
            "open_at": today.isoformat(),
            "close_at": (today + timezone.timedelta(days=30)).isoformat(),
            "is_active": "on",
        }
        response = self.client.post(reverse("manage_section_forms"), payload, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(FormTemplate.objects.filter(code="smme-q2", section=self.section).exists())
        self.assertContains(response, "Created form SMME Q2")

    def test_section_admin_updates_schedule(self):
        self.client.force_login(self.section_admin)
        open_at = timezone.now().date() - timezone.timedelta(days=1)
        close_at = timezone.now().date() + timezone.timedelta(days=14)
        payload = {
            "action": "update_schedule",
            "form_id": self.form.id,
            "open_at": open_at.isoformat(),
            "close_at": close_at.isoformat(),
            "is_active": "on",
        }
        response = self.client.post(reverse("manage_section_forms"), payload, follow=True)
        self.assertEqual(response.status_code, 200)
        self.form.refresh_from_db()
        self.assertEqual(self.form.open_at, open_at)
        self.assertEqual(self.form.close_at, close_at)
        self.assertTrue(self.form.is_active)
        self.assertContains(response, "Updated schedule for")

    def test_cannot_manage_other_section_forms(self):
        other_form = FormTemplate.objects.create(
            section=self.other_section,
            code="hrd-q1",
            title="HRD Q1",
            period_type=FormTemplate.PeriodType.QUARTER,
            open_at=timezone.now().date(),
            close_at=timezone.now().date(),
        )
        self.client.force_login(self.section_admin)
        payload = {
            "action": "close_today",
            "form_id": other_form.id,
        }
        response = self.client.post(reverse("manage_section_forms"), payload)
        self.assertEqual(response.status_code, 404)


class Form1ModelValidationTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.section = Section.objects.create(code="validation", name="Validation")
        self.school = School.objects.create(code="validation-school", name="Validation School")
        today = timezone.now().date()
        self.period = Period.objects.create(
            label="Q2",
            school_year_start=2025,
            quarter_tag='Q2',
            display_order=2,
            is_active=True,
        )
        self.form = FormTemplate.objects.create(
            section=self.section,
            code="validation-form",
            title="Validation Form",
            open_at=today,
            close_at=today,
        )
        self.user = User.objects.create_user(username="validator", password="pass123")
        self.submission = Submission.objects.create(
            school=self.school,
            form_template=self.form,
            period=self.period,
            submitted_by=self.user,
        )

    def test_slp_row_total_not_exceed_enrolment(self):
        row = Form1SLPRow(
            submission=self.submission,
            grade_label="Grade 1",
            enrolment=10,
            dnme=2,
            fs=2,
            s=2,
            vs=2,
            o=3,
        )
        with self.assertRaises(ValidationError):
            row.clean()

    def test_rma_row_total_not_exceed_enrolment(self):
        row = Form1RMARow(
            submission=self.submission,
            grade_label="g1",
            enrolment=20,
            band_below_75=5,
            band_75_79=5,
            band_80_84=5,
            band_85_89=5,
            band_90_100=1,
        )
        with self.assertRaises(ValidationError):
            row.clean()






